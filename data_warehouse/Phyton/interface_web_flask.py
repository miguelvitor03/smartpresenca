from flask import Flask, request, render_template_string, send_file, make_response, redirect, url_for, session
import mysql.connector
import csv
import xml.dom.minidom as md
import xml.etree.ElementTree as ET
from datetime import datetime
import os
import io
import tempfile
from functools import wraps

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

import openpyxl

app = Flask(__name__)
app.secret_key = 'chave_secreta_para_sessao'  # Necessária para usar sessões

# Template para o login
LOGIN_TEMPLATE = """
<!doctype html>
<html>
<head>
    <title>Login - Sistema de Relatório de Presença</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
            line-height: 1.6;
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
            background-color: #f5f5f5;
        }
        .login-container {
            background-color: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
            width: 350px;
        }
        h1 {
            color: #333366;
            text-align: center;
            margin-bottom: 20px;
        }
        form {
            display: flex;
            flex-direction: column;
        }
        .form-group {
            margin-bottom: 15px;
        }
        label {
            font-weight: bold;
            margin-bottom: 5px;
            display: block;
        }
        input[type="text"], input[type="password"] {
            padding: 10px;
            width: 100%;
            border: 1px solid #ddd;
            border-radius: 4px;
            box-sizing: border-box;
        }
        input[type="submit"] {
            background-color: #FFBF00;
            color: white;
            padding: 12px;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-weight: bold;
            margin-top: 10px;
        }
        input[type="submit"]:hover {
            background-color: #f1a700;
        }
        .error {
            color: red;
            font-weight: bold;
            text-align: center;
            margin-bottom: 15px;
        }
    </style>
</head>
<body>
    <div class="login-container">
        <h1>Sistema de Presença</h1>
        
        {% if error %}
        <div class="error">{{ error }}</div>
        {% endif %}
        
        <form method="POST" action="/login">
            <div class="form-group">
                <label for="ra">RA:</label>
                <input type="text" name="ra" id="ra" required>
            </div>
            <div class="form-group">
                <label for="senha">Senha:</label>
                <input type="password" name="senha" id="senha" required>
            </div>
            <input type="submit" value="Entrar">
        </form>
    </div>
</body>
</html>
"""

# Template principal para consulta de presença
HTML_TEMPLATE = """
<!doctype html>
<html>
<head>
    <title>Relatório de Presença</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
            line-height: 1.6;
        }
        h1 {
            color: #333366;
            border-bottom: 2px solid #eee;
            padding-bottom: 10px;
        }
        .user-info {
            background-color: #f9f9f9;
            padding: 10px;
            border-radius: 5px;
            margin-bottom: 20px;
        }
        .logout-btn {
            background-color: #ff6b6b;
            color: white;
            padding: 5px 10px;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            text-decoration: none;
            display: inline-block;
            font-size: 14px;
        }
        .logout-btn:hover {
            background-color: #e55c5c;
        }
        form {
            background-color: #f5f5f5;
            padding: 15px;
            border-radius: 5px;
            margin-bottom: 20px;
        }
        input[type="text"] {
            padding: 8px;
            width: 200px;
            border: 1px solid #ddd;
            border-radius: 4px;
        }
        input[type="submit"], button {
            background-color: #FFBF00;
            color: white;
            padding: 8px 15px;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            margin-right: 5px;
        }
        input[type="submit"]:hover, button:hover {
            background-color: #f1a700;
        }
        table {
            border-collapse: collapse;
            width: 100%;
            margin-top: 20px;
        }
        th, td {
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }
        th {
            background-color: #333366;
            color: white;
        }
        tr:nth-child(even) {
            background-color: #f2f2f2;
        }
        .export-options {
            margin-top: 20px;
            padding: 10px;
            background-color: #f9f9f9;
            border-radius: 5px;
        }
        .error {
            color: red;
            font-weight: bold;
        }
        .success {
            color: green;
            font-weight: bold;
        }
    </style>
</head>
<body>
    <div class="user-info">
        <p>Logado como: <strong>{{ aluno_nome }}</strong> (RA: {{ ra_aluno }})</p>
        <a href="/logout" class="logout-btn">Sair</a>
    </div>
    
    <h1>Consulta de Presença</h1>
    
    {% if resultados %}
        <h2>Seu histórico de presenças</h2>
        <table>
            <tr>
                <th>Disciplina</th>
                <th>Data da Aula</th>
                <th>Presença</th>
            </tr>
            {% for linha in resultados %}
            <tr>
                <td>{{ linha[0] }}</td>
                <td>{{ linha[1] }}</td>
                <td>{{ linha[2] }}</td>
            </tr>
            {% endfor %}
        </table>
        
        <div class="export-options">
            <h3>Exportar relatório</h3>
            <form action="/export" method="post">
                <button type="submit" name="format" value="csv">CSV</button>
                <button type="submit" name="format" value="xml">XML</button>
                <button type="submit" name="format" value="pdf">PDF</button>
                <button type="submit" name="format" value="excel">Excel</button>
            </form>
        </div>
    {% else %}
        <p class="error">Nenhum registro de presença encontrado para seu RA.</p>
    {% endif %}
</body>
</html>
"""

# Função para conectar ao banco de dados
def conectar_bd():
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="Luhar2923",
            database="Banco_pi"  # Nome do banco de dados atualizado
        )
        return conn
    except mysql.connector.Error as err:
        print(f"Erro ao conectar ao MySQL: {err}")
        return None

# Função para verificar login
def verificar_login(ra, senha):
    conn = conectar_bd()
    if not conn:
        return None
    
    cursor = conn.cursor()
    
    # Verificar se o RA e senha coincidem na tabela de alunos
    query = "SELECT id_aluno, nome FROM aluno WHERE ra = %s AND senha = %s"
    cursor.execute(query, (ra, senha))
    resultado = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return resultado

# Decorator para verificar se o usuário está logado
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Função para buscar dados do aluno
def buscar_dados_aluno(id_aluno):
    conn = conectar_bd()
    if not conn:
        return None, None, None
    
    cursor = conn.cursor()
    
    # Primeiro, obter informações básicas do aluno
    aluno_query = "SELECT nome, ra FROM aluno WHERE id_aluno = %s"
    cursor.execute(aluno_query, (id_aluno,))
    aluno_info = cursor.fetchone()
    
    if not aluno_info:
        cursor.close()
        conn.close()
        return None, None, None
    
    aluno_nome, ra = aluno_info
    
    # Buscar registros de presença
    query = """
    SELECT 
        d.nome AS disciplina,
        t.data_aula,
        p.presente
    FROM presenca p
    JOIN disciplina d ON d.iddisciplina = p.id_disciplina
    JOIN tempo t ON t.id_tempo = p.id_tempo
    WHERE p.id_aluno = %s
    ORDER BY t.data_aula;
    """
    
    cursor.execute(query, (id_aluno,))
    dados = cursor.fetchall()
    
    resultados = []
    
    for row in dados:
        disciplina = row[0]
        data_aula = row[1].strftime("%Y-%m-%d")
        status = "Presente" if row[2] else "Ausente"
        resultados.append((disciplina, data_aula, status))
    
    cursor.close()
    conn.close()
    
    return aluno_nome, ra, resultados

# Função para gerar CSV
def gerar_csv(aluno_nome, ra, resultados):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Aluno', 'RA', 'Disciplina', 'Data da Aula', 'Presença'])
    
    for disciplina, data_aula, status in resultados:
        writer.writerow([aluno_nome, ra, disciplina, data_aula, status])
    
    output.seek(0)
    return output

# Função para gerar XML
def gerar_xml(aluno_nome, ra, resultados):
    # Criar a estrutura XML
    root = ET.Element("RelatorioPresenca")
    aluno_element = ET.SubElement(root, "Aluno", RA=str(ra), Nome=aluno_nome)
    
    # Adicionar registros de presença
    for disciplina, data_aula, status in resultados:
        presenca = ET.SubElement(aluno_element, "Registro")
        ET.SubElement(presenca, "Disciplina").text = disciplina
        ET.SubElement(presenca, "DataAula").text = data_aula
        ET.SubElement(presenca, "Status").text = status
    
    # Salvar o XML formatado
    xmlstr = ET.tostring(root, encoding='utf-8')
    dom = md.parseString(xmlstr)
    pretty_xml = dom.toprettyxml(indent="  ")
    
    return io.BytesIO(pretty_xml.encode('utf-8'))

# Função para gerar PDF
def gerar_pdf(aluno_nome, ra, resultados):
    buffer = io.BytesIO()
    
    # Criar PDF
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    
    # Título
    elements.append(Paragraph(f"Relatório de Presença - {aluno_nome} (RA: {ra})", title_style))
    
    # Dados para a tabela
    data = [['Aluno', 'Disciplina', 'Data da Aula', 'Presença']]
    
    for disciplina, data_aula, status in resultados:
        data.append([aluno_nome, disciplina, data_aula, status])
    
    # Criar tabela
    table = Table(data)
    
    # Estilo da tabela
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    
    table.setStyle(style)
    elements.append(table)
    
    # Construir o PDF
    doc.build(elements)
    
    buffer.seek(0)
    return buffer

# Função para gerar Excel
def gerar_excel(aluno_nome, ra, resultados):
    output = io.BytesIO()
    
    # Criar uma nova planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Presença RA {ra}"
    
    # Adicionar cabeçalho
    ws.append(['Aluno', 'RA', 'Disciplina', 'Data da Aula', 'Presença'])
    
    # Estilizar cabeçalho
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Adicionar dados
    for disciplina, data_aula, status in resultados:
        ws.append([aluno_nome, ra, disciplina, data_aula, status])
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar o arquivo
    wb.save(output)
    output.seek(0)
    return output

@app.route('/', methods=['GET'])
def index():
    if 'user_id' in session:
        return redirect(url_for('consulta'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    
    if request.method == 'POST':
        ra = request.form['ra']
        senha = request.form['senha']
        
        resultado = verificar_login(ra, senha)
        
        if resultado:
            user_id, nome = resultado
            session['user_id'] = user_id
            session['nome'] = nome
            session['ra'] = ra
            return redirect(url_for('consulta'))
        else:
            error = 'RA ou senha inválidos. Por favor, tente novamente.'
    
    return render_template_string(LOGIN_TEMPLATE, error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/consulta')
@login_required
def consulta():
    user_id = session['user_id']
    
    aluno_nome, ra, resultados = buscar_dados_aluno(user_id)
    
    return render_template_string(HTML_TEMPLATE, resultados=resultados, aluno_nome=aluno_nome, ra_aluno=ra)

@app.route('/export', methods=['POST'])
@login_required
def export():
    user_id = session['user_id']
    formato = request.form['format']
    
    aluno_nome, ra, resultados = buscar_dados_aluno(user_id)
    
    if not resultados:
        return "Nenhum dado encontrado para exportação", 404
    
    if formato == 'csv':
        output = gerar_csv(aluno_nome, ra, resultados)
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=relatorio_presenca_RA_{ra}.csv"
        response.headers["Content-type"] = "text/csv"
        return response
        
    elif formato == 'xml':
        output = gerar_xml(aluno_nome, ra, resultados)
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=relatorio_presenca_RA_{ra}.xml"
        response.headers["Content-type"] = "application/xml"
        return response
        
    elif formato == 'pdf':
        output = gerar_pdf(aluno_nome, ra, resultados)
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=relatorio_presenca_RA_{ra}.pdf"
        response.headers["Content-type"] = "application/pdf"
        return response
        
    elif formato == 'excel':
        output = gerar_excel(aluno_nome, ra, resultados)
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=relatorio_presenca_RA_{ra}.xlsx"
        response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response
    
    return "Formato não suportado", 400

if __name__ == '__main__':
    app.run(debug=True)


#http://localhost:5000 