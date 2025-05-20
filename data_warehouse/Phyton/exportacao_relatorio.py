import mysql.connector
import csv
import xml.dom.minidom as md
import xml.etree.ElementTree as ET
from datetime import datetime
import os
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl

# Conectar ao banco MySQL
def conectar_bd():
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="Luhar2923",
            database="data_pi"
        )
        return conn
    except mysql.connector.Error as err:
        print(f"Erro ao conectar ao MySQL: {err}")
        exit(1)

# Função para buscar dados do aluno
def buscar_dados_aluno(ra):
    conn = conectar_bd()
    cursor = conn.cursor()
    
    query = """
    SELECT 
        a.nome AS aluno_nome,
        d.nome AS disciplina,
        t.data_aula,
        f.presente
    FROM fato_presenca f
    JOIN dim_aluno a ON a.id_aluno = f.id_aluno
    JOIN dim_disciplina d ON d.iddisciplina = f.id_disciplina
    JOIN dim_tempo t ON t.id_tempo = f.id_tempo
    WHERE a.ra = %s
    ORDER BY t.data_aula;
    """
    
    cursor.execute(query, (ra,))
    resultados = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return resultados

# Função para gerar relatório CSV
def gerar_csv(ra, resultados):
    if not resultados:
        return False
        
    aluno_nome = resultados[0][0]
    nome_arquivo = f"relatorio_presenca_RA_{ra}.csv"
    
    with open(nome_arquivo, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(['Aluno', 'Disciplina', 'Data da Aula', 'Presença'])
        
        for row in resultados:
            aluno_nome, disciplina, data_aula, presente = row
            status = "Presente" if presente else "Ausente"
            writer.writerow([aluno_nome, disciplina, data_aula.strftime("%Y-%m-%d"), status])
    
    print(f"✅ Relatório CSV gerado com sucesso: {nome_arquivo}")
    return nome_arquivo

# Função para gerar relatório XML
def gerar_xml(ra, resultados):
    if not resultados:
        return False
        
    aluno_nome = resultados[0][0]
    nome_arquivo = f"relatorio_presenca_RA_{ra}.xml"
    
    # Criar a estrutura XML
    root = ET.Element("RelatorioPresenca")
    aluno_element = ET.SubElement(root, "Aluno", RA=ra, Nome=aluno_nome)
    
    # Adicionar registros de presença
    for row in resultados:
        aluno_nome, disciplina, data_aula, presente = row
        status = "Presente" if presente else "Ausente"
        
        presenca = ET.SubElement(aluno_element, "Registro")
        ET.SubElement(presenca, "Disciplina").text = disciplina
        ET.SubElement(presenca, "DataAula").text = data_aula.strftime("%Y-%m-%d")
        ET.SubElement(presenca, "Status").text = status
    
    # Salvar o XML formatado
    xmlstr = ET.tostring(root, encoding='utf-8')
    dom = md.parseString(xmlstr)
    pretty_xml = dom.toprettyxml(indent="  ")
    
    with open(nome_arquivo, "w", encoding='utf-8') as f:
        f.write(pretty_xml)
    
    print(f"✅ Relatório XML gerado com sucesso: {nome_arquivo}")
    return nome_arquivo

# Função para gerar relatório PDF
def gerar_pdf(ra, resultados):
    if not resultados:
        return False
        
    aluno_nome = resultados[0][0]
    nome_arquivo = f"relatorio_presenca_RA_{ra}.pdf"
    
    # Criar PDF
    doc = SimpleDocTemplate(nome_arquivo, pagesize=letter)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    
    # Título
    elements.append(Paragraph(f"Relatório de Presença - {aluno_nome} (RA: {ra})", title_style))
    
    # Dados para a tabela
    data = [['Aluno', 'Disciplina', 'Data da Aula', 'Presença']]
    
    for row in resultados:
        aluno_nome, disciplina, data_aula, presente = row
        status = "Presente" if presente else "Ausente"
        data.append([aluno_nome, disciplina, data_aula.strftime("%Y-%m-%d"), status])
    
    # Criar tabela
    table = Table(data)
    
    # Estilo da tabela
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    
    table.setStyle(style)
    elements.append(table)
    
    # Construir o PDF
    doc.build(elements)
    
    print(f"✅ Relatório PDF gerado com sucesso: {nome_arquivo}")
    return nome_arquivo

# Função para gerar relatório Excel
def gerar_excel(ra, resultados):
    if not resultados:
        return False
        
    aluno_nome = resultados[0][0]
    nome_arquivo = f"relatorio_presenca_RA_{ra}.xlsx"
    
    # Criar uma nova planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Presença RA {ra}"
    
    # Adicionar cabeçalho
    ws.append(['Aluno', 'Disciplina', 'Data da Aula', 'Presença'])
    
    # Estilizar cabeçalho
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Adicionar dados
    for row in resultados:
        aluno_nome, disciplina, data_aula, presente = row
        status = "Presente" if presente else "Ausente"
        ws.append([aluno_nome, disciplina, data_aula.strftime("%Y-%m-%d"), status])
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar o arquivo
    wb.save(nome_arquivo)
    
    print(f"✅ Relatório Excel gerado com sucesso: {nome_arquivo}")
    return nome_arquivo

# Função principal
def main():
    # Solicitar o RA do aluno
    ra_input = input("Digite o RA do aluno: ")
    
    # Buscar dados
    resultados = buscar_dados_aluno(ra_input)
    
    if not resultados:
        print("Nenhum registro encontrado para este RA.")
        return
    
    continuar_gerando = True
    
    while continuar_gerando:
        # Menu de opções
        print("\nEscolha o formato do relatório:")
        print("1 - CSV")
        print("2 - XML")
        print("3 - PDF")
        print("4 - Excel")
        
        opcao = input("Digite o número da opção desejada: ")
        
        arquivo_gerado = None
        
        if opcao == "1":
            arquivo_gerado = gerar_csv(ra_input, resultados)
        elif opcao == "2":
            arquivo_gerado = gerar_xml(ra_input, resultados)
        elif opcao == "3":
            arquivo_gerado = gerar_pdf(ra_input, resultados)
        elif opcao == "4":
            arquivo_gerado = gerar_excel(ra_input, resultados)
        else:
            print("Opção inválida!")
            continue
        
        if arquivo_gerado:
            # Perguntar se deseja gerar outro tipo de relatório
            resposta = input("\nDeseja gerar outro tipo de relatório para o mesmo RA? (s/n): ").lower()
            continuar_gerando = resposta == 's' or resposta == 'sim'
        else:
            continuar_gerando = False
    
    print("\nPrograma finalizado. Obrigado por utilizar o sistema de relatórios!")

# Executar o programa
if __name__ == "__main__":
    main()